#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
MEMRISTORES - create_master_workbook.py
Versión: 2.0.0

Genera:
1. data/dictionaries/MEMRISTORES_Clinical_Radiological_Dictionary_v2.0.xlsx
2. data/dictionaries/MEMRISTORES_Variable_Dictionary_v2.0.csv
3. data/templates/MEMRISTORES_MasterDataset_Template_v2.0.csv
4. data/processed/MEMRISTORES_MasterDataset_v2.0.xlsx

Uso:
    python src/create_master_workbook.py

Dependencia:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "Falta la dependencia 'openpyxl'. Instálala con:\n"
        "    pip install openpyxl"
    ) from exc


VERSION = "2.0.0"
DICTIONARY_VERSION = "MEMRISTORES-DD-2.0"
TODAY = date.today().isoformat()


@dataclass(frozen=True)
class VariableDefinition:
    variable: str
    domain: str
    data_type: str
    allowed_values: str
    operational_definition: str
    required_sequence: str
    coding_rule: str
    source_standard: str
    priority: str = "Core"


VARIABLES: tuple[VariableDefinition, ...] = (
    # Identificación y trazabilidad
    VariableDefinition(
        "patient_id", "identification", "string", "ACV-0001…",
        "Identificador pseudonimizado estable por paciente.",
        "clinical registry",
        "Asignación ascendente. No reutilizar códigos.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "study_id", "identification", "string", "ACV-0001-RM01…",
        "Identificador único por estudio de imagen.",
        "clinical registry",
        "Mantener patient_id y numerar estudios en orden cronológico.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "cohort_year", "clinical", "integer", "2022…",
        "Año de inclusión en la cohorte.",
        "clinical registry",
        "Usar únicamente el año, no la fecha completa.",
        "MEMRISTORES governance", "Core"
    ),
    VariableDefinition(
        "index_event_type", "clinical", "categorical",
        "ischemic_stroke|TIA|ICH|unknown",
        "Tipo de evento cerebrovascular índice.",
        "clinical registry",
        "No inferir si no consta.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "age_at_event", "clinical", "integer", "0-120|NA",
        "Edad en años al momento del evento índice.",
        "clinical registry",
        "No publicar fecha de nacimiento.",
        "MEMRISTORES governance", "Core"
    ),
    VariableDefinition(
        "sex", "clinical", "categorical", "F|M|Other_NR|NA",
        "Sexo registrado para análisis descriptivo.",
        "clinical registry",
        "Usar únicamente categorías previamente definidas.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "days_event_to_imaging", "temporal", "integer", "0…|NA",
        "Número de días entre el evento índice y la imagen.",
        "clinical registry",
        "Preferir diferencia calculada; evitar fechas exactas en datos públicos.",
        "MEMRISTORES governance", "Critical"
    ),

    # Adquisición
    VariableDefinition(
        "modality", "acquisition", "categorical",
        "MRI|CT|CTA|MRA|Other",
        "Modalidad principal del estudio.",
        "study metadata",
        "No inferir.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "field_strength_T", "acquisition", "numeric",
        "1.5|3.0|Other|NA",
        "Intensidad de campo de la resonancia.",
        "study metadata",
        "NA cuando no conste.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "image_quality", "quality", "categorical",
        "adequate|limited|non_evaluable|NA",
        "Calidad global para lectura de biomarcadores.",
        "all available sequences",
        "Registrar la causa de limitación en notes.",
        "MEMRISTORES QC", "Critical"
    ),
    VariableDefinition(
        "sequence_T1", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de secuencia T1.",
        "study metadata",
        "1 presente; 0 evaluada y ausente; NS no realizada/no disponible.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "sequence_T2", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de secuencia T2.",
        "study metadata",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "sequence_FLAIR", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de FLAIR.",
        "study metadata",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "sequence_DWI", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de DWI.",
        "study metadata",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "sequence_ADC", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de mapa ADC.",
        "study metadata",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "sequence_GRE_SWI", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de T2*-GRE o SWI.",
        "study metadata",
        "Sin esta secuencia, microbleeds y siderosis deben registrarse como NS.",
        "MARS/BOMBS", "Critical"
    ),
    VariableDefinition(
        "sequence_MRA_CTA", "sequence_availability", "binary_missing",
        "0|1|NA|NE|NS",
        "Disponibilidad de angiografía por RM o TC.",
        "study metadata",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Core"
    ),

    # Lesión índice
    VariableDefinition(
        "acute_infarct_present", "index_lesion", "binary_missing",
        "0|1|NA|NE|NS",
        "Lesión isquémica aguda o reciente compatible con el evento índice.",
        "DWI+ADC±FLAIR",
        "No usar 0 si faltan DWI/ADC y el informe no permite excluirla.",
        "STRIVE-2", "Critical"
    ),
    VariableDefinition(
        "acute_infarct_count", "index_lesion", "integer",
        "0…|NA|NE",
        "Número de lesiones agudas separadas.",
        "DWI+ADC",
        "Contar focos anatómicamente separados.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "acute_infarct_laterality", "index_lesion", "categorical",
        "right|left|bilateral|midline|NA|NE",
        "Lateralidad de la lesión índice.",
        "DWI+ADC±FLAIR",
        "Usar bilateral cuando existan lesiones en ambos hemisferios.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "acute_infarct_territory", "index_lesion", "categorical_multivalue",
        "ACA|MCA|PCA|vertebrobasilar|borderzone|perforator|other|NA|NE",
        "Territorio vascular dominante.",
        "DWI+ADC±MRA/CTA",
        "Separar múltiples territorios con |.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "acute_infarct_location_detail", "index_lesion", "text",
        "controlled free text",
        "Localización anatómica detallada.",
        "DWI+ADC±FLAIR",
        "Registrar estructura, lado y profundidad.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "acute_infarct_max_diameter_mm", "index_lesion", "numeric",
        "mm|NA|NE",
        "Diámetro máximo de la lesión aguda.",
        "DWI or FLAIR",
        "Medir en el plano de mayor extensión.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "acute_infarct_volume_ml", "index_lesion", "numeric",
        "mL|NA|NE",
        "Volumen estimado de la lesión aguda.",
        "DWI or FLAIR",
        "Registrar método de segmentación en notes.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "dwi_positive", "index_lesion", "binary_missing",
        "0|1|NA|NE|NS",
        "Hiperintensidad compatible con restricción en DWI.",
        "DWI",
        "Evaluar junto con ADC.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "adc_restriction", "index_lesion", "binary_missing",
        "0|1|NA|NE|NS",
        "Disminución del ADC correspondiente.",
        "ADC",
        "No concluir restricción verdadera con DWI aislada.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "flair_positive", "index_lesion", "binary_missing",
        "0|1|NA|NE|NS",
        "Hiperintensidad FLAIR correspondiente a la lesión índice.",
        "FLAIR",
        "Aplicar leyenda estandarizada.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "dwi_flair_mismatch", "index_lesion", "binary_missing",
        "0|1|NA|NE|NS",
        "DWI positiva sin hiperintensidad FLAIR correspondiente.",
        "DWI+FLAIR",
        "Solo evaluable si ambas secuencias existen y son interpretables.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "hemorrhagic_transformation_present", "complication", "binary_missing",
        "0|1|NA|NE|NS",
        "Transformación hemorrágica de una lesión isquémica.",
        "CT or GRE/SWI",
        "No confundir con hemorragia remota.",
        "ECASS", "Core"
    ),
    VariableDefinition(
        "hemorrhagic_transformation_ECASS", "complication", "categorical",
        "HI1|HI2|PH1|PH2|NA|NE",
        "Clasificación radiológica ECASS.",
        "CT or GRE/SWI",
        "Codificar únicamente cuando existan elementos suficientes.",
        "ECASS", "Core"
    ),

    # Enfermedad de pequeños vasos
    VariableDefinition(
        "lacune_present", "CSVD", "binary_missing",
        "0|1|NA|NE|NS",
        "Cavidad subcortical compatible con infarto o hemorragia antigua de pequeño vaso.",
        "T1+T2+FLAIR",
        "Diferenciar de espacios perivasculares.",
        "STRIVE-2", "Critical"
    ),
    VariableDefinition(
        "lacune_count", "CSVD", "integer",
        "0…|NA|NE",
        "Número total de lacunas.",
        "T1+T2+FLAIR",
        "Contar cavidades separadas.",
        "STRIVE-2", "Core"
    ),
    VariableDefinition(
        "lacune_locations", "CSVD", "categorical_multivalue",
        "basal_ganglia|thalamus|internal_capsule|corona_radiata|brainstem|cerebellum|other",
        "Distribución anatómica de lacunas.",
        "T1+T2+FLAIR",
        "Separar localizaciones con |.",
        "STRIVE-2", "Core"
    ),
    VariableDefinition(
        "recent_small_subcortical_infarct", "CSVD", "binary_missing",
        "0|1|NA|NE|NS",
        "Infarto reciente pequeño en territorio de arteriola perforante.",
        "DWI+ADC±FLAIR",
        "Registrar independientemente de lacuna antigua.",
        "STRIVE-2", "Core"
    ),
    VariableDefinition(
        "WMH_present", "WMH", "binary_missing",
        "0|1|NA|NE|NS",
        "Hiperintensidades de sustancia blanca de probable origen vascular.",
        "FLAIR or T2",
        "Excluir edema y otras causas cuando sea posible.",
        "STRIVE-2", "Critical"
    ),
    VariableDefinition(
        "fazekas_periventricular", "WMH", "ordinal",
        "0|1|2|3|NA|NE|NS",
        "Carga visual periventricular de WMH.",
        "FLAIR",
        "Aplicar criterios Fazekas definidos en la hoja Scales.",
        "Fazekas", "Critical"
    ),
    VariableDefinition(
        "fazekas_deep", "WMH", "ordinal",
        "0|1|2|3|NA|NE|NS",
        "Carga visual profunda de WMH.",
        "FLAIR",
        "Aplicar criterios Fazekas definidos en la hoja Scales.",
        "Fazekas", "Critical"
    ),
    VariableDefinition(
        "microbleed_present", "microbleeds", "binary_missing",
        "0|1|NA|NE|NS",
        "Microhemorragias cerebrales crónicas.",
        "GRE/SWI",
        "Sin GRE/SWI, registrar NS.",
        "MARS/BOMBS", "Critical"
    ),
    VariableDefinition(
        "microbleed_total_count", "microbleeds", "integer",
        "0…|NA|NE|NS",
        "Número total de microbleeds.",
        "GRE/SWI",
        "Excluir vasos, calcificaciones y artefactos.",
        "MARS/BOMBS", "Core"
    ),
    VariableDefinition(
        "microbleed_lobar_count", "microbleeds", "integer",
        "0…|NA|NE|NS",
        "Conteo lobar.",
        "GRE/SWI",
        "Clasificar por topografía MARS.",
        "MARS", "Core"
    ),
    VariableDefinition(
        "microbleed_deep_count", "microbleeds", "integer",
        "0…|NA|NE|NS",
        "Conteo profundo.",
        "GRE/SWI",
        "Clasificar por topografía MARS.",
        "MARS", "Core"
    ),
    VariableDefinition(
        "microbleed_infratentorial_count", "microbleeds", "integer",
        "0…|NA|NE|NS",
        "Conteo infratentorial.",
        "GRE/SWI",
        "Clasificar por topografía MARS.",
        "MARS", "Core"
    ),
    VariableDefinition(
        "microbleed_distribution", "microbleeds", "categorical",
        "none|strictly_lobar|deep|infratentorial|mixed|NA|NE|NS",
        "Distribución anatómica global de microbleeds.",
        "GRE/SWI",
        "Derivar del conteo por compartimentos.",
        "MARS", "Critical"
    ),
    VariableDefinition(
        "cortical_superficial_siderosis", "hemorrhagic_markers", "binary_missing",
        "0|1|NA|NE|NS",
        "Depósito lineal de hemosiderina en surcos corticales.",
        "GRE/SWI",
        "Diferenciar de vasos y artefactos.",
        "STRIVE-2", "Core"
    ),
    VariableDefinition(
        "PVS_basal_ganglia_grade", "PVS", "ordinal",
        "0|1|2|3|4|NA|NE|NS",
        "Carga visual de espacios perivasculares en ganglios basales.",
        "T2",
        "Usar la misma escala y corte de referencia en toda la cohorte.",
        "STRIVE-2", "Core"
    ),
    VariableDefinition(
        "PVS_centrum_semiovale_grade", "PVS", "ordinal",
        "0|1|2|3|4|NA|NE|NS",
        "Carga visual de espacios perivasculares en centrum semiovale.",
        "T2",
        "Usar la misma escala y corte de referencia en toda la cohorte.",
        "STRIVE-2", "Core"
    ),

    # Atrofia y lesiones crónicas
    VariableDefinition(
        "brain_atrophy_present", "atrophy", "binary_missing",
        "0|1|NA|NE|NS",
        "Atrofia cerebral visualmente apreciable.",
        "T1 or FLAIR or CT",
        "Registrar escala específica cuando sea posible.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "global_cortical_atrophy_grade", "atrophy", "ordinal",
        "0|1|2|3|NA|NE|NS",
        "Grado visual global de atrofia cortical.",
        "T1 or FLAIR or CT",
        "Aplicar una escala consistente.",
        "GCA visual scale", "Optional"
    ),
    VariableDefinition(
        "medial_temporal_atrophy_right", "atrophy", "ordinal",
        "0|1|2|3|4|NA|NE|NS",
        "Atrofia temporal medial derecha.",
        "coronal T1",
        "Aplicar escala MTA de forma consistente.",
        "Scheltens MTA", "Optional"
    ),
    VariableDefinition(
        "medial_temporal_atrophy_left", "atrophy", "ordinal",
        "0|1|2|3|4|NA|NE|NS",
        "Atrofia temporal medial izquierda.",
        "coronal T1",
        "Aplicar escala MTA de forma consistente.",
        "Scheltens MTA", "Optional"
    ),
    VariableDefinition(
        "strategic_infarct_present", "chronic_lesions", "binary_missing",
        "0|1|NA|NE|NS",
        "Infarto en localización potencialmente estratégica para cognición o conducta.",
        "T1+T2+FLAIR±DWI",
        "Registrar localización; no atribuir causalidad automática.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "strategic_infarct_location", "chronic_lesions", "text",
        "controlled free text",
        "Localización anatómica del infarto estratégico.",
        "T1+T2+FLAIR±DWI",
        "Ejemplos: tálamo, hipocampo, giro angular, caudado, territorio ACA.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "old_cortical_infarct_present", "chronic_lesions", "binary_missing",
        "0|1|NA|NE|NS",
        "Infarto cortical crónico.",
        "T1+T2+FLAIR or CT",
        "Registrar encefalomalacia/gliosis asociada.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "old_subcortical_infarct_present", "chronic_lesions", "binary_missing",
        "0|1|NA|NE|NS",
        "Infarto subcortical crónico.",
        "T1+T2+FLAIR or CT",
        "Diferenciar de lacunas.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "old_infratentorial_infarct_present", "chronic_lesions", "binary_missing",
        "0|1|NA|NE|NS",
        "Infarto crónico de tronco o cerebelo.",
        "T1+T2+FLAIR or CT",
        "Registrar localización.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "encephalomalacia_gliosis", "chronic_lesions", "binary_missing",
        "0|1|NA|NE|NS",
        "Encefalomalacia o gliosis residual.",
        "T1+T2+FLAIR or CT",
        "Registrar distribución en notes.",
        "MEMRISTORES protocol", "Core"
    ),

    # Vasculatura
    VariableDefinition(
        "large_artery_stenosis_present", "vasculature", "binary_missing",
        "0|1|NA|NE|NS",
        "Estenosis arterial intracraneal o extracraneal relevante.",
        "CTA or MRA",
        "No concluir ausencia sin estudio vascular.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "stenosis_vessel", "vasculature", "categorical_multivalue",
        "ICA|MCA|ACA|PCA|VA|BA|other|NA|NE|NS",
        "Vaso o vasos con estenosis.",
        "CTA or MRA",
        "Separar múltiples vasos con |.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "stenosis_percent_category", "vasculature", "categorical",
        "none|mild|moderate|severe|occlusion|NA|NE|NS",
        "Categoría de severidad de estenosis.",
        "CTA or MRA",
        "Documentar método si se dispone.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "intracranial_atherosclerosis", "vasculature", "binary_missing",
        "0|1|NA|NE|NS",
        "Aterosclerosis intracraneal.",
        "CTA or MRA",
        "No inferir con imagen parenquimatosa aislada.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "dolichoectasia", "vasculature", "binary_missing",
        "0|1|NA|NE|NS",
        "Dolicoectasia arterial.",
        "CTA or MRA",
        "Registrar vaso y método en notes.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "aneurysm_present", "vasculature", "binary_missing",
        "0|1|NA|NE|NS",
        "Aneurisma intracraneal.",
        "CTA or MRA",
        "Registrar localización y tamaño en notes.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "vascular_malformation_present", "vasculature", "binary_missing",
        "0|1|NA|NE|NS",
        "Malformación vascular intracraneal.",
        "CTA or MRA±contrast MRI",
        "Registrar tipo en notes.",
        "MEMRISTORES protocol", "Optional"
    ),

    # Escalas derivadas y resumen
    VariableDefinition(
        "total_CSVD_score_0_4", "derived_score", "integer",
        "0|1|2|3|4|NA",
        "Puntuación total pragmática de enfermedad cerebral de pequeños vasos.",
        "multisequence MRI",
        "1 punto por lacuna, microbleed, WMH moderada-grave y PVS-BG moderados-graves.",
        "Staals et al.", "Critical"
    ),
    VariableDefinition(
        "probable_CAA_Boston2_category", "derived_pattern", "categorical",
        "not_applicable|not_met|possible|probable|NA",
        "Categoría exploratoria según criterios Boston 2.0.",
        "GRE/SWI plus clinical data",
        "No calcular con imagen aislada incompleta.",
        "Boston criteria v2.0", "Exploratory"
    ),
    VariableDefinition(
        "vascular_parkinsonism_pattern", "derived_pattern", "categorical",
        "none|compatible|indeterminate|NA",
        "Patrón radiológico compatible con parkinsonismo vascular.",
        "MRI",
        "No equivale a diagnóstico clínico.",
        "MEMRISTORES exploratory", "Exploratory"
    ),
    VariableDefinition(
        "other_incidental_findings", "other", "text",
        "free text",
        "Hallazgos incidentales no cubiertos por otras variables.",
        "all available sequences",
        "Evitar datos identificables.",
        "MEMRISTORES protocol", "Optional"
    ),
    VariableDefinition(
        "radiology_text_deidentified", "traceability", "text",
        "free text",
        "Extracto desidentificado del informe radiológico.",
        "radiology report",
        "Eliminar nombres, IDs, fechas completas, direcciones y teléfonos.",
        "MEMRISTORES governance", "Critical"
    ),
    VariableDefinition(
        "coder_summary", "traceability", "text",
        "free text",
        "Resumen estructurado del codificador.",
        "all available sources",
        "Separar hallazgos, negativos evaluables y limitaciones.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "coding_confidence", "quality", "categorical",
        "high|moderate|low",
        "Confianza global de la codificación.",
        "all available sources",
        "Alta: hallazgos claros y secuencias adecuadas.",
        "MEMRISTORES QC", "Critical"
    ),
    VariableDefinition(
        "primary_reader", "quality", "string",
        "initials_or_role",
        "Lector primario.",
        "coding log",
        "No usar nombre completo en repositorios públicos.",
        "MEMRISTORES QC", "Core"
    ),
    VariableDefinition(
        "second_reader", "quality", "string",
        "initials_or_role|NA",
        "Segundo lector.",
        "coding log",
        "NA si no se realizó doble lectura.",
        "MEMRISTORES QC", "Core"
    ),
    VariableDefinition(
        "adjudicated", "quality", "binary_missing",
        "0|1|NA",
        "Indica si una discrepancia fue adjudicada.",
        "coding log",
        "1 solo después de resolución documentada.",
        "MEMRISTORES QC", "Core"
    ),
    VariableDefinition(
        "dictionary_version", "traceability", "string",
        DICTIONARY_VERSION,
        "Versión del diccionario aplicada.",
        "coding log",
        "No modificar retrospectivamente sin registrar migración.",
        "MEMRISTORES protocol", "Critical"
    ),
    VariableDefinition(
        "coding_date", "traceability", "date",
        "YYYY-MM-DD",
        "Fecha de codificación.",
        "coding log",
        "Fecha de trabajo, no fecha clínica.",
        "MEMRISTORES protocol", "Core"
    ),
    VariableDefinition(
        "source_type", "traceability", "categorical",
        "report_only|images_only|report_and_images",
        "Fuente usada para codificar.",
        "coding log",
        "No afirmar revisión de imágenes cuando solo se usó el informe.",
        "MEMRISTORES QC", "Critical"
    ),
    VariableDefinition(
        "notes", "traceability", "text",
        "free text",
        "Observaciones metodológicas.",
        "all available sources",
        "No incluir información identificable.",
        "MEMRISTORES governance", "Core"
    ),
)


MASTER_COLUMNS: tuple[str, ...] = tuple(variable.variable for variable in VARIABLES)


SCALE_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("Missingness", "0", "Evaluado y ausente.", "No equivale a NA, NE o NS."),
    ("Missingness", "1", "Presente.", "Debe existir trazabilidad."),
    ("Missingness", "NA", "Dato no disponible.", "No consta o no fue proporcionado."),
    ("Missingness", "NE", "No evaluable.", "La secuencia existe, pero la calidad impide decidir."),
    ("Missingness", "NS", "Secuencia necesaria no realizada/no disponible.", "No debe convertirse en 0."),
    ("Fazekas periventricular", "0", "Ausente.", "FLAIR."),
    ("Fazekas periventricular", "1", "Casquetes o línea fina.", "Leve."),
    ("Fazekas periventricular", "2", "Halo liso.", "Moderada."),
    ("Fazekas periventricular", "3", "Extensión irregular a sustancia blanca profunda.", "Grave."),
    ("Fazekas deep", "0", "Ausente.", "FLAIR."),
    ("Fazekas deep", "1", "Focos puntiformes.", "Leve."),
    ("Fazekas deep", "2", "Confluencia inicial.", "Moderada."),
    ("Fazekas deep", "3", "Grandes áreas confluentes.", "Grave."),
    ("Total CSVD", "Lacune", "1 punto si existe ≥1 lacuna.", "Definición STRIVE-2."),
    ("Total CSVD", "Microbleed", "1 punto si existe ≥1 microbleed.", "Requiere GRE/SWI."),
    ("Total CSVD", "WMH", "1 punto si Fazekas deep 2–3 o PV 3.", "Regla pragmática."),
    ("Total CSVD", "PVS-BG", "1 punto si PVS-BG grado 2–4.", "Mantener la misma escala."),
    ("ECASS HT", "HI1", "Pequeñas petequias periféricas sin efecto de masa.", ""),
    ("ECASS HT", "HI2", "Petequias confluentes sin efecto de masa.", ""),
    ("ECASS HT", "PH1", "Hematoma ≤30% del área infartada con leve efecto de masa.", ""),
    ("ECASS HT", "PH2", "Hematoma >30% con efecto de masa significativo.", ""),
)


SOURCES: tuple[tuple[str, str, str], ...] = (
    (
        "STRIVE-2",
        "Duering M, et al. Neuroimaging standards for research into small vessel disease—advances since 2013. Lancet Neurol. 2023.",
        "Definiciones y reporte de marcadores de enfermedad cerebral de pequeños vasos."
    ),
    (
        "Total CSVD score",
        "Staals J, et al. Stroke subtype, vascular risk factors, and total MRI brain small-vessel disease burden. Neurology. 2014.",
        "Puntuación total pragmática 0–4."
    ),
    (
        "MARS",
        "Gregoire SM, et al. The Microbleed Anatomical Rating Scale. Neurology. 2009.",
        "Definición y distribución anatómica de microbleeds."
    ),
    (
        "BOMBS",
        "Cordonnier C, et al. Improving interrater agreement about brain microbleeds. Stroke. 2009.",
        "Identificación estandarizada de microbleeds."
    ),
    (
        "Fazekas",
        "Fazekas F, et al. MR signal abnormalities at 1.5 T in Alzheimer's dementia and normal aging. AJR. 1987.",
        "Gradación visual de WMH."
    ),
    (
        "ECASS",
        "Clasificación ECASS de transformación hemorrágica.",
        "Categorías HI1, HI2, PH1 y PH2."
    ),
)


THIN_GRAY = Side(style="thin", color="D9E1F2")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
TITLE_FILL = PatternFill("solid", fgColor="17365D")


def ensure_directories(root: Path) -> dict[str, Path]:
    paths = {
        "dictionaries": root / "data" / "dictionaries",
        "templates": root / "data" / "templates",
        "processed": root / "data" / "processed",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)


def add_table(ws, ref: str, display_name: str) -> None:
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_rows(ws, rows: Iterable[Iterable[object]], start_row: int = 1) -> int:
    current = start_row
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=current, column=col_idx, value=value)
        current += 1
    return current - 1


def create_dictionary_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"

    ws.merge_cells("A1:H1")
    ws["A1"] = "MEMRISTORES CLINICAL-RADIOLOGICAL DICTIONARY"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    readme_rows = (
        ("Field", "Value"),
        ("Project", "MEMRISTORES"),
        ("Dictionary version", DICTIONARY_VERSION),
        ("Script version", VERSION),
        ("Generated", TODAY),
        ("Unit of analysis", "One row per imaging study"),
        ("Patient coding", "ACV-0001, ACV-0002, ..."),
        ("Study coding", "ACV-0001-RM01, ACV-0001-RM02, ..."),
        ("Missing values", "0=absent; 1=present; NA=not available; NE=not evaluable; NS=sequence unavailable"),
        ("Governance", "Do not upload identifiers, full reports, DICOM or exact clinical dates to a public repository."),
    )
    last = write_rows(ws, readme_rows, start_row=3)
    style_header(ws, 3)
    set_widths(ws, {"A": 26, "B": 100})
    for row in ws.iter_rows(min_row=4, max_row=last, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    variables_ws = wb.create_sheet("Variables")
    variable_headers = (
        "variable", "domain", "data_type", "allowed_values",
        "operational_definition", "required_sequence", "coding_rule",
        "source_standard", "priority"
    )
    variables_ws.append(variable_headers)
    for item in VARIABLES:
        variables_ws.append((
            item.variable,
            item.domain,
            item.data_type,
            item.allowed_values,
            item.operational_definition,
            item.required_sequence,
            item.coding_rule,
            item.source_standard,
            item.priority,
        ))
    style_header(variables_ws)
    variables_ws.freeze_panes = "A2"
    set_widths(variables_ws, {
        "A": 34, "B": 22, "C": 20, "D": 40, "E": 62,
        "F": 28, "G": 65, "H": 28, "I": 15,
    })
    for row in variables_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    add_table(variables_ws, f"A1:I{variables_ws.max_row}", "VariableDictionary")

    scales_ws = wb.create_sheet("Scales")
    scales_ws.append(("scale_marker", "category", "operational_criterion", "methodological_note"))
    for row in SCALE_ROWS:
        scales_ws.append(row)
    style_header(scales_ws)
    scales_ws.freeze_panes = "A2"
    set_widths(scales_ws, {"A": 28, "B": 20, "C": 72, "D": 52})
    for row in scales_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    add_table(scales_ws, f"A1:D{scales_ws.max_row}", "ScaleLegend")

    sources_ws = wb.create_sheet("Sources")
    sources_ws.append(("standard", "reference", "use_in_MEMRISTORES"))
    for row in SOURCES:
        sources_ws.append(row)
    style_header(sources_ws)
    set_widths(sources_ws, {"A": 24, "B": 90, "C": 70})
    for row in sources_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    add_table(sources_ws, f"A1:C{sources_ws.max_row}", "SourcesTable")

    history_ws = wb.create_sheet("Version_History")
    history_ws.append(("version", "date", "change", "author_or_role"))
    history_ws.append((VERSION, TODAY, "Initial version 2.0 modular workbook and dataset generator.", "MEMRISTORES team"))
    style_header(history_ws)
    set_widths(history_ws, {"A": 15, "B": 15, "C": 85, "D": 28})
    add_table(history_ws, f"A1:D{history_ws.max_row}", "VersionHistory")

    wb.save(path)


def create_csv_dictionary(path: Path) -> None:
    headers = (
        "variable", "domain", "data_type", "allowed_values",
        "operational_definition", "required_sequence", "coding_rule",
        "source_standard", "priority"
    )
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(headers)
        for item in VARIABLES:
            writer.writerow((
                item.variable,
                item.domain,
                item.data_type,
                item.allowed_values,
                item.operational_definition,
                item.required_sequence,
                item.coding_rule,
                item.source_standard,
                item.priority,
            ))


def create_master_csv(path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(MASTER_COLUMNS)


def create_master_workbook(path: Path, max_rows: int = 500) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER_DATASET"

    for col_idx, column_name in enumerate(MASTER_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=column_name)

    style_header(ws)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(MASTER_COLUMNS)).coordinate}"

    for col_idx, column_name in enumerate(MASTER_COLUMNS, start=1):
        width = 18
        if column_name in {
            "acute_infarct_location_detail",
            "lacune_locations",
            "strategic_infarct_location",
            "other_incidental_findings",
            "radiology_text_deidentified",
            "coder_summary",
            "notes",
        }:
            width = 42
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Validaciones comunes
    binary_missing = DataValidation(
        type="list",
        formula1='"0,1,NA,NE,NS"',
        allow_blank=True,
        error="Use uno de estos valores: 0, 1, NA, NE o NS.",
        errorTitle="Valor no permitido",
    )
    ws.add_data_validation(binary_missing)

    sex_validation = DataValidation(type="list", formula1='"F,M,Other_NR,NA"', allow_blank=True)
    ws.add_data_validation(sex_validation)

    modality_validation = DataValidation(type="list", formula1='"MRI,CT,CTA,MRA,Other"', allow_blank=True)
    ws.add_data_validation(modality_validation)

    quality_validation = DataValidation(
        type="list",
        formula1='"adequate,limited,non_evaluable,NA"',
        allow_blank=True,
    )
    ws.add_data_validation(quality_validation)

    confidence_validation = DataValidation(type="list", formula1='"high,moderate,low"', allow_blank=True)
    ws.add_data_validation(confidence_validation)

    source_validation = DataValidation(
        type="list",
        formula1='"report_only,images_only,report_and_images"',
        allow_blank=True,
    )
    ws.add_data_validation(source_validation)

    ecass_validation = DataValidation(type="list", formula1='"HI1,HI2,PH1,PH2,NA,NE"', allow_blank=True)
    ws.add_data_validation(ecass_validation)

    fazekas_validation = DataValidation(type="list", formula1='"0,1,2,3,NA,NE,NS"', allow_blank=True)
    ws.add_data_validation(fazekas_validation)

    pvs_validation = DataValidation(type="list", formula1='"0,1,2,3,4,NA,NE,NS"', allow_blank=True)
    ws.add_data_validation(pvs_validation)

    column_index = {name: idx + 1 for idx, name in enumerate(MASTER_COLUMNS)}

    for variable in VARIABLES:
        col_letter = ws.cell(row=1, column=column_index[variable.variable]).column_letter
        target = f"{col_letter}2:{col_letter}{max_rows + 1}"

        if variable.data_type == "binary_missing":
            binary_missing.add(target)
        elif variable.variable == "sex":
            sex_validation.add(target)
        elif variable.variable == "modality":
            modality_validation.add(target)
        elif variable.variable == "image_quality":
            quality_validation.add(target)
        elif variable.variable == "coding_confidence":
            confidence_validation.add(target)
        elif variable.variable == "source_type":
            source_validation.add(target)
        elif variable.variable == "hemorrhagic_transformation_ECASS":
            ecass_validation.add(target)
        elif variable.variable in {"fazekas_periventricular", "fazekas_deep"}:
            fazekas_validation.add(target)
        elif variable.variable in {"PVS_basal_ganglia_grade", "PVS_centrum_semiovale_grade"}:
            pvs_validation.add(target)

    # Fórmula automática para Total CSVD
    csfd_col = column_index["total_CSVD_score_0_4"]
    lacune_col = column_index["lacune_present"]
    microbleed_col = column_index["microbleed_present"]
    fazekas_pv_col = column_index["fazekas_periventricular"]
    fazekas_deep_col = column_index["fazekas_deep"]
    pvs_bg_col = column_index["PVS_basal_ganglia_grade"]

    for row in range(2, max_rows + 2):
        lac = ws.cell(row=row, column=lacune_col).coordinate
        cmb = ws.cell(row=row, column=microbleed_col).coordinate
        fpv = ws.cell(row=row, column=fazekas_pv_col).coordinate
        fdp = ws.cell(row=row, column=fazekas_deep_col).coordinate
        pvs = ws.cell(row=row, column=pvs_bg_col).coordinate
        target = ws.cell(row=row, column=csfd_col)
        target.value = (
            f'=IF(OR({lac}="NA",{lac}="NE",{lac}="NS",'
            f'{cmb}="NA",{cmb}="NE",{cmb}="NS",'
            f'{fpv}="NA",{fpv}="NE",{fpv}="NS",'
            f'{fdp}="NA",{fdp}="NE",{fdp}="NS",'
            f'{pvs}="NA",{pvs}="NE",{pvs}="NS"),'
            f'"NA",'
            f'IF({lac}=1,1,0)+'
            f'IF({cmb}=1,1,0)+'
            f'IF(OR({fdp}=2,{fdp}=3,{fpv}=3),1,0)+'
            f'IF(OR({pvs}=2,{pvs}=3,{pvs}=4),1,0))'
        )

    # Versión y fecha de codificación prellenadas
    dict_version_col = column_index["dictionary_version"]
    coding_date_col = column_index["coding_date"]
    for row in range(2, max_rows + 2):
        ws.cell(row=row, column=dict_version_col, value=DICTIONARY_VERSION)
        ws.cell(row=row, column=coding_date_col, value=TODAY)

    # Resaltado básico de problemas de calidad
    ns_fill = PatternFill("solid", fgColor="FFF2CC")
    ne_fill = PatternFill("solid", fgColor="F4CCCC")
    end_cell = ws.cell(row=max_rows + 1, column=len(MASTER_COLUMNS)).coordinate
    ws.conditional_formatting.add(
        f"A2:{end_cell}",
        FormulaRule(formula=['A2="NS"'], fill=ns_fill)
    )
    ws.conditional_formatting.add(
        f"A2:{end_cell}",
        FormulaRule(formula=['A2="NE"'], fill=ne_fill)
    )

    # Hojas auxiliares
    legend_ws = wb.create_sheet("LEGEND")
    legend_ws.append(("code", "meaning"))
    legend_ws.append(("0", "Evaluado y ausente"))
    legend_ws.append(("1", "Presente"))
    legend_ws.append(("NA", "Dato no disponible"))
    legend_ws.append(("NE", "No evaluable"))
    legend_ws.append(("NS", "Secuencia necesaria no realizada/no disponible"))
    style_header(legend_ws)
    set_widths(legend_ws, {"A": 16, "B": 75})
    add_table(legend_ws, f"A1:B{legend_ws.max_row}", "MissingnessLegend")

    summary_ws = wb.create_sheet("PATIENT_SUMMARIES")
    summary_ws.append((
        "patient_id", "study_id", "neuroradiological_summary",
        "key_biomarkers", "limitations", "review_status"
    ))
    style_header(summary_ws)
    set_widths(summary_ws, {"A": 18, "B": 20, "C": 70, "D": 55, "E": 50, "F": 22})
    status_validation = DataValidation(
        type="list",
        formula1='"pending,coded,second_read,adjudicated"',
        allow_blank=True,
    )
    summary_ws.add_data_validation(status_validation)
    status_validation.add(f"F2:F{max_rows + 1}")
    summary_ws.freeze_panes = "A2"

    add_table(
        ws,
        f"A1:{ws.cell(row=max_rows + 1, column=len(MASTER_COLUMNS)).coordinate}",
        "MasterDataset"
    )

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate MEMRISTORES dictionary and master dataset files."
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Repository root. Default: parent directory of src/.",
    )
    parser.add_argument(
        "--rows",
        type=int,
        default=500,
        help="Number of empty study rows in the Excel master dataset.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = args.root.resolve()

    if args.rows < 1:
        print("Error: --rows debe ser mayor o igual a 1.", file=sys.stderr)
        return 2

    paths = ensure_directories(root)

    dictionary_xlsx = paths["dictionaries"] / "MEMRISTORES_Clinical_Radiological_Dictionary_v2.0.xlsx"
    dictionary_csv = paths["dictionaries"] / "MEMRISTORES_Variable_Dictionary_v2.0.csv"
    master_csv = paths["templates"] / "MEMRISTORES_MasterDataset_Template_v2.0.csv"
    master_xlsx = paths["processed"] / "MEMRISTORES_MasterDataset_v2.0.xlsx"

    create_dictionary_workbook(dictionary_xlsx)
    create_csv_dictionary(dictionary_csv)
    create_master_csv(master_csv)
    create_master_workbook(master_xlsx, max_rows=args.rows)

    print("\nMEMRISTORES files generated successfully:")
    for output in (dictionary_xlsx, dictionary_csv, master_csv, master_xlsx):
        print(f" - {output.relative_to(root)}")

    print(
        "\nImportant: review governance requirements before uploading any patient-derived data."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
